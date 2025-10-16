import os
from django.core.management.base import BaseCommand
from django.apps import apps
import openpyxl
from openpyxl.utils import get_column_letter
from django.conf import settings

class Command(BaseCommand):
    help = "Backup all database tables to a single Excel file (db_backup.xlsx)"

    def handle(self, *args, **options):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet
        models = apps.get_models()
        for model in models:
            name = model._meta.db_table
            ws = wb.create_sheet(title=name[:31])  # Excel sheet name max 31 chars
            fields = [f.name for f in model._meta.fields]
            ws.append(fields)
            for obj in model.objects.all():
                row = [str(getattr(obj, f)) if getattr(obj, f) is not None else '' for f in fields]
                ws.append(row)
            # Autosize columns
            for col in ws.columns:
                max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = max(10, min(max_length+2, 40))
        backup_path = os.path.join(settings.BASE_DIR, 'db_backup.xlsx')
        wb.save(backup_path)
        self.stdout.write(self.style.SUCCESS(f"Backup complete: {backup_path}"))
