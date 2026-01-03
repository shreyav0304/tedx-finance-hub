from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.forms import UserCreationForm
from django.contrib import messages
from django.db.models import Sum, Q
from django.http import HttpResponse, JsonResponse
from django.core.exceptions import ValidationError
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.core.cache import cache
from datetime import datetime, timedelta
from django.db.models.functions import TruncMonth
import logging

from .models import ManagementFund, Sponsor, Transaction, Category, UserPreference
from .forms import (
    TransactionForm,
    ManagementFundForm,
    SponsorForm,
    UserPreferenceForm,
    validate_file_extension,
    validate_file_size,
)

import openpyxl

# Configure logger
logger = logging.getLogger(__name__)


# --- Helper functions ---
def is_in_group(user, group_name):
    """
    Checks if a user is in a given group.
    Admins (staff users) automatically have access to 'Treasurer' group.
    """
    if user.is_authenticated:
        # Admins get automatic access to Treasurer group
        if group_name == 'Treasurer' and user.is_staff:
            return True
        return user.groups.filter(name=group_name).exists()
    return False


def parse_date(date_str):
    """
    Safely parse a date string in YYYY-MM-DD format.
    Returns a date object or None if parsing fails.
    """
    if date_str:
        try:
            return datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            return None
    return None


def get_cached_category_choices():
    """Return merged category choices (dynamic + defaults) with short caching."""
    cache_key = 'tedx_category_choices'
    cached = cache.get(cache_key)
    if cached:
        return cached

    try:
        dynamic = list(Category.objects.all().values_list('name', 'name'))
    except Exception:
        dynamic = []
    default_choices = list(getattr(Transaction, 'CATEGORY_CHOICES', []))
    seen = set()
    merged = []
    for val, label in dynamic + default_choices:
        if val not in seen:
            merged.append((val, label))
            seen.add(val)

    cache.set(cache_key, merged, 300)  # 5 minutes
    return merged


def invalidate_category_cache():
    cache.delete('tedx_category_choices')


def get_sponsor_tier(amount):
    """
    Returns sponsor tier name and badge class based on amount.
    Thresholds (INR): Gold ≥ 200,000; Silver ≥ 50,000; Bronze < 50,000
    """
    try:
        amt = float(amount)
    except (ValueError, TypeError):
        amt = 0.0
    
    if amt >= 200000:
        return 'Gold', 'bg-yellow-500 text-black'
    elif amt >= 50000:
        return 'Silver', 'bg-slate-300 text-slate-900'
    else:
        return 'Bronze', 'bg-amber-700 text-white'


def apply_transaction_filters(request, queryset=None, user_is_treasurer=False):
    """
    Apply comprehensive filters to transaction queryset based on request parameters.
    Reusable for both table views and exports.
    
    Args:
        request: HTTP request with filter parameters
        queryset: Initial queryset (defaults to all transactions)
        user_is_treasurer: Whether to apply treasurer-only filters
        
    Returns:
        Filtered queryset
    """
    if queryset is None:
        queryset = Transaction.objects.all().select_related('created_by')
    else:
        queryset = queryset.select_related('created_by')
    
    # Search filter (searches in title, category, description)
    search_query = request.GET.get('search', '').strip()
    if search_query:
        queryset = queryset.filter(
            Q(title__icontains=search_query) |
            Q(category__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    # Status filter
    status_filter = request.GET.get('status', 'all')
    if status_filter == 'approved':
        queryset = queryset.filter(approved=True)
    elif status_filter == 'pending':
        queryset = queryset.filter(approved=False)
    
    # Category filter
    category_filter = request.GET.get('category', 'all')
    if category_filter != 'all':
        queryset = queryset.filter(category=category_filter)
    
    # Date range filters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date:
        queryset = queryset.filter(date__gte=start_date)
    if end_date:
        queryset = queryset.filter(date__lte=end_date)
    
    # Amount range filters
    min_amount = request.GET.get('min_amount')
    max_amount = request.GET.get('max_amount')
    if min_amount:
        queryset = queryset.filter(amount__gte=min_amount)
    if max_amount:
        queryset = queryset.filter(amount__lte=max_amount)
    
    # Submitted by filter (for treasurers only)
    submitted_by = request.GET.get('submitted_by')
    if submitted_by and user_is_treasurer:
        queryset = queryset.filter(created_by__username__icontains=submitted_by)
    
    return queryset

# --- Authentication ---
def signup(request):
    """User signup with email verification."""
    from .models import EmailVerification
    from .utils import generate_email_token, send_verification_email
    
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            # Create user but mark as inactive until email verified
            user = form.save(commit=False)
            user.is_active = False  # Deactivate until email verified
            user.save()
            
            # Create email verification record
            token = generate_email_token()
            EmailVerification.objects.create(user=user, token=token)
            
            # Send verification email
            if send_verification_email(user, token, request):
                messages.success(
                    request,
                    'Account created! Please check your email to verify your account. '
                    'Verification link expires in 24 hours.'
                )
                return redirect('login')
            else:
                # If email fails, delete the user and show error
                user.delete()
                messages.error(request, 'Failed to send verification email. Please try again.')
    else:
        form = UserCreationForm()
    
    context = {'form': form}
    return render(request, 'registration/signup.html', context)


def verify_email(request, token):
    """Verify user email from token."""
    from .models import EmailVerification
    from django.utils import timezone
    
    try:
        verification = EmailVerification.objects.get(token=token)
        
        # Check if token is still valid (24 hours)
        if not verification.is_token_valid(hours=24):
            messages.error(request, 'Verification link has expired. Please sign up again.')
            return redirect('signup')
        
        # Mark email as verified and activate user
        verification.is_verified = True
        verification.verified_at = timezone.now()
        verification.save()
        
        user = verification.user
        user.is_active = True
        user.save()
        
        messages.success(request, 'Email verified successfully! You can now log in.')
        return redirect('login')
    
    except EmailVerification.DoesNotExist:
        messages.error(request, 'Invalid verification link.')
        return redirect('signup')


def login_view(request):
    """Login view with rate limiting."""
    from django.contrib.auth import authenticate, login as auth_login
    from .models import LoginAttempt
    from .utils import get_client_ip
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        ip_address = get_client_ip(request)
        
        # Check if IP is rate limited
        if LoginAttempt.is_rate_limited(ip_address, max_attempts=5, time_window=300):
            messages.error(
                request,
                'Too many failed login attempts. Please try again in 5 minutes.'
            )
            return render(request, 'registration/login.html')
        
        # Attempt authentication
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            # Successful login
            LoginAttempt.log_attempt(username, ip_address, success=True)
            
            # Check if email is verified
            if hasattr(user, 'email_verification') and not user.email_verification.is_verified:
                messages.warning(request, 'Please verify your email before logging in.')
                return redirect('login')
            
            auth_login(request, user)
            messages.success(request, f'Welcome back, {user.first_name or username}!')
            return redirect('dashboard')
        else:
            # Failed login
            LoginAttempt.log_attempt(username, ip_address, success=False)
            messages.error(request, 'Invalid username or password.')
    
    return render(request, 'registration/login.html')

# --- Core Views ---
@login_required
def budgets(request):
    """Budget tracking view showing all budgets and their utilization."""
    from .models import Budget
    user_is_treasurer = is_in_group(request.user, 'Treasurer')
    budgets = Budget.objects.all().order_by('category__name')
    context = {
        'budgets': budgets,
        'is_treasurer': user_is_treasurer,
    }
    return render(request, 'tedx_finance/budgets.html', context)


@login_required
def budget_suggestions(request):
    """
    AI-powered budget prediction and recommendations.
    
    Analyzes:
    - Historical spending patterns
    - Burn rate (spending velocity)
    - Category-wise allocation efficiency
    - Projected budget needs for upcoming period
    - Risk of budget overrun
    """
    from .models import Budget
    from django.db.models import Sum, Count, Avg
    from django.db.models.functions import TruncWeek
    from datetime import datetime, timedelta
    import json
    
    user_is_treasurer = is_in_group(request.user, 'Treasurer')
    
    # Get all budgets and current financial state
    budgets = Budget.objects.all().order_by('category__name')
    
    # Calculate total income
    total_income = (
        (ManagementFund.objects.aggregate(total=Sum('amount'))['total'] or 0) +
        (Sponsor.objects.aggregate(total=Sum('amount'))['total'] or 0)
    )
    
    # Calculate total spent (approved transactions)
    total_spent_val = Transaction.objects.filter(
        approved=True,
        amount__lt=0
    ).aggregate(total=Sum('amount'))['total'] or 0
    total_spent = abs(total_spent_val)
    
    remaining_funds = total_income - total_spent
    
    # Analyze spending patterns (last 30 days)
    thirty_days_ago = datetime.now().date() - timedelta(days=30)
    recent_spending = abs(Transaction.objects.filter(
        approved=True,
        amount__lt=0,
        date__gte=thirty_days_ago
    ).aggregate(total=Sum('amount'))['total'] or 0)
    
    # Calculate burn rate (spending per day)
    days_analyzed = min(30, (datetime.now().date() - thirty_days_ago).days)
    if days_analyzed > 0:
        daily_burn_rate = recent_spending / days_analyzed
        weekly_burn_rate = daily_burn_rate * 7
        monthly_burn_rate = daily_burn_rate * 30
    else:
        daily_burn_rate = weekly_burn_rate = monthly_burn_rate = 0
    
    # Calculate runway (days until funds run out at current burn rate)
    if daily_burn_rate > 0:
        runway_days = int(remaining_funds / daily_burn_rate)
    else:
        runway_days = 999  # Infinite if not spending
    
    # Category-wise analysis with predictions
    category_insights = []
    total_suggested_budget = 0
    
    for budget in budgets:
        spent = budget.spent()
        remaining = budget.remaining()
        utilization = budget.utilization_percent()
        
        # Calculate category burn rate (last 30 days)
        category_recent_spending = abs(Transaction.objects.filter(
            category=budget.category.name,
            approved=True,
            amount__lt=0,
            date__gte=thirty_days_ago
        ).aggregate(total=Sum('amount'))['total'] or 0)
        
        if days_analyzed > 0:
            category_daily_burn = category_recent_spending / days_analyzed
        else:
            category_daily_burn = 0
        
        # Calculate days remaining in budget period
        days_remaining = max(0, (budget.end_date - datetime.now().date()).days)
        
        # Predict total spending by end of budget period
        predicted_additional_spend = category_daily_burn * days_remaining
        predicted_total_spend = spent + predicted_additional_spend
        
        # Calculate suggested budget (120% of predicted to have buffer)
        suggested_budget = predicted_total_spend * 1.2
        total_suggested_budget += suggested_budget
        
        # Determine status and risk level
        if utilization >= 100:
            status = 'exceeded'
            risk_level = 'critical'
        elif utilization >= 80:
            status = 'warning'
            risk_level = 'high'
        elif utilization >= 60:
            status = 'moderate'
            risk_level = 'medium'
        else:
            status = 'good'
            risk_level = 'low'
        
        # Calculate if more budget is needed
        budget_shortage = max(0, predicted_total_spend - float(budget.amount))
        
        category_insights.append({
            'category': budget.category.name,
            'category_code': budget.category.name,
            'current_budget': float(budget.amount),
            'spent': spent,
            'remaining': remaining,
            'utilization': round(utilization, 1),
            'daily_burn_rate': round(category_daily_burn, 2),
            'predicted_total_spend': round(predicted_total_spend, 2),
            'suggested_budget': round(suggested_budget, 2),
            'budget_shortage': round(budget_shortage, 2),
            'days_remaining': days_remaining,
            'status': status,
            'risk_level': risk_level,
            'needs_increase': budget_shortage > 0,
        })
    
    # Overall recommendations
    overall_budget_needed = max(total_suggested_budget, total_spent + (monthly_burn_rate * 2))
    additional_funds_needed = max(0, overall_budget_needed - total_income)
    
    # Generate insights and recommendations
    recommendations = []
    
    if runway_days < 30:
        recommendations.append({
            'type': 'critical',
            'icon': '🚨',
            'title': 'Critical: Funds Running Low',
            'message': f'At current spending rate, you have only {runway_days} days of runway. Immediate action required!',
            'action': f'Secure additional ₹{round(monthly_burn_rate * 2):,.0f} to ensure 60 days of operations.'
        })
    elif runway_days < 60:
        recommendations.append({
            'type': 'warning',
            'icon': '⚠️',
            'title': 'Warning: Limited Runway',
            'message': f'You have {runway_days} days of runway remaining.',
            'action': f'Plan to secure ₹{round(monthly_burn_rate):,.0f} within next month.'
        })
    
    for insight in category_insights:
        if insight['needs_increase']:
            recommendations.append({
                'type': 'info',
                'icon': '💡',
                'title': f"{insight['category']} Budget Increase Needed",
                'message': f"Current trend suggests you'll exceed budget by ₹{insight['budget_shortage']:,.0f}",
                'action': f"Consider increasing {insight['category']} budget to ₹{insight['suggested_budget']:,.0f}"
            })
    
    if additional_funds_needed > 0:
        recommendations.append({
            'type': 'info',
            'icon': '💰',
            'title': 'Additional Funding Required',
            'message': f'To complete event safely, you need ₹{round(additional_funds_needed):,.0f} more',
            'action': 'Focus on sponsor acquisition or reduce discretionary spending'
        })
    
    if not recommendations:
        recommendations.append({
            'type': 'success',
            'icon': '✅',
            'title': 'Budget Healthy',
            'message': 'Your budget allocation looks good! All categories are on track.',
            'action': 'Continue monitoring spending patterns regularly.'
        })
    
    # Weekly spending trend (last 8 weeks)
    eight_weeks_ago = datetime.now().date() - timedelta(weeks=8)
    weekly_spending = Transaction.objects.filter(
        approved=True,
        amount__lt=0,
        date__gte=eight_weeks_ago
    ).annotate(
        week=TruncWeek('date')
    ).values('week').annotate(
        total=Sum('amount')
    ).order_by('week')
    
    spending_trend_labels = [item['week'].strftime('%b %d') for item in weekly_spending]
    spending_trend_values = [abs(float(item['total'])) for item in weekly_spending]
    
    context = {
        'is_treasurer': user_is_treasurer,
        'total_income': total_income,
        'total_spent': total_spent,
        'remaining_funds': remaining_funds,
        'daily_burn_rate': round(daily_burn_rate, 2),
        'weekly_burn_rate': round(weekly_burn_rate, 2),
        'monthly_burn_rate': round(monthly_burn_rate, 2),
        'runway_days': runway_days,
        'category_insights': category_insights,
        'recommendations': recommendations,
        'total_suggested_budget': round(total_suggested_budget, 2),
        'additional_funds_needed': round(additional_funds_needed, 2),
        'spending_trend_labels': json.dumps(spending_trend_labels),
        'spending_trend_values': json.dumps(spending_trend_values),
    }
    
    return render(request, 'tedx_finance/budget_suggestions.html', context)
@login_required
def transactions_table(request):
    """Excel-like table view with inline editing capabilities and advanced filtering"""
    user_is_treasurer = is_in_group(request.user, 'Treasurer')
    
    # Apply filters using helper function
    transactions = apply_transaction_filters(request, user_is_treasurer=user_is_treasurer)
    
    # Order by (default: newest first)
    order_by = request.GET.get('order_by', '-date')
    valid_order_fields = ['date', '-date', 'amount', '-amount', 'title', '-title', 'category', '-category']
    if order_by in valid_order_fields:
        transactions = transactions.order_by(order_by)
    else:
        transactions = transactions.order_by('-date')
    
    # Dynamic categories merged with defaults for filters
    categories = get_cached_category_choices()

    context = {
        'transactions': transactions,
        'categories': categories,
        'is_treasurer': user_is_treasurer,
        'search_query': request.GET.get('search', ''),
        'status_filter': request.GET.get('status', 'all'),
        'category_filter': request.GET.get('category', 'all'),
        'start_date': request.GET.get('start_date'),
        'end_date': request.GET.get('end_date'),
        'min_amount': request.GET.get('min_amount'),
        'max_amount': request.GET.get('max_amount'),
        'submitted_by': request.GET.get('submitted_by'),
        'order_by': order_by,
    }
    return render(request, 'tedx_finance/transactions_table.html', context)

@login_required
def dashboard(request):
    """
    Dashboard view with analytics, charts, recent transactions, and sponsor tiers.
    Supports optional date range filtering via GET parameters.
    """
    context = {}
    try:
        user_is_treasurer = is_in_group(request.user, 'Treasurer')

        # Parse optional date range filters
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        start_date = parse_date(start_date_str)
        end_date = parse_date(end_date_str)

        # Income aggregates (respect date range if provided)
        mf_qs = ManagementFund.objects.all()
        sp_qs = Sponsor.objects.all()
        if start_date:
            mf_qs = mf_qs.filter(date_received__gte=start_date)
            sp_qs = sp_qs.filter(date_received__gte=start_date)
        if end_date:
            mf_qs = mf_qs.filter(date_received__lte=end_date)
            sp_qs = sp_qs.filter(date_received__lte=end_date)

        management_funds = mf_qs.aggregate(total=Sum('amount'))['total'] or 0
        sponsor_funds = sp_qs.aggregate(total=Sum('amount'))['total'] or 0
        total_funds = management_funds + sponsor_funds

        # Approved transactions with optional date filters
        approved_transactions = Transaction.objects.filter(approved=True)
        if start_date:
            approved_transactions = approved_transactions.filter(date__gte=start_date)
        if end_date:
            approved_transactions = approved_transactions.filter(date__lte=end_date)

        total_spent_val = approved_transactions.filter(amount__lt=0).aggregate(total=Sum('amount'))['total'] or 0
        total_spent = abs(total_spent_val)
        remaining_balance = total_funds - total_spent

        # Chart Data - Category spending (expenses only)
        category_spending = list(
            approved_transactions.filter(amount__lt=0)
            .values('category')
            .annotate(total=Sum('amount'))
            .order_by('category')
        )

        # Convert category codes to human labels
        category_map = dict(getattr(Transaction, 'CATEGORY_CHOICES', []))
        for item in category_spending:
            # If not in default map, keep the raw category string (user-defined)
            item['category'] = category_map.get(item['category'], item['category'])

        income_data = [float(management_funds), float(sponsor_funds)]

        # Recent and Pending Transactions
        transactions = approved_transactions.order_by('-date')[:10]
        pending_transactions = Transaction.objects.filter(approved=False).select_related('created_by').order_by('-date')

        # Sponsor tiers (within selected date range)
        sponsors_qs = sp_qs.order_by('-amount')
        sponsors_with_tiers = []
        for s in sponsors_qs:
            tier, badge_class = get_sponsor_tier(s.amount)
            sponsors_with_tiers.append({
                'name': s.name,
                'amount': float(s.amount),
                'tier': tier,
                'badge_class': badge_class,
            })

        # Spending Trends: within selected range, else last 6 months
        if start_date or end_date:
            trend_qs = approved_transactions.filter(amount__lt=0)
        else:
            six_months_ago = datetime.now().date() - timedelta(days=180)
            trend_qs = approved_transactions.filter(amount__lt=0, date__gte=six_months_ago)

        monthly_spending = (
            trend_qs
            .annotate(month=TruncMonth('date'))
            .values('month')
            .annotate(total=Sum('amount'))
            .order_by('month')
        )

        spending_months = [item['month'].strftime('%b %Y') for item in monthly_spending]
        spending_amounts = [abs(float(item['total'])) for item in monthly_spending]

        # Budget vs Actual Analysis
        from .models import Budget
        budgets = Budget.objects.all().order_by('category__name')
        
        budget_comparison = []
        total_budget_amount = 0
        total_budget_spent = 0
        budget_exceeded_count = 0
        budget_warning_count = 0
        
        for budget in budgets:
            spent = budget.spent()
            remaining = budget.remaining()
            utilization = budget.utilization_percent()
            
            # Categorize budget status
            if budget.is_exceeded():
                status = 'exceeded'
                budget_exceeded_count += 1
            elif utilization >= 80:
                status = 'warning'
                budget_warning_count += 1
            else:
                status = 'healthy'
            
            budget_comparison.append({
                'category': budget.category.name,
                'budget_amount': float(budget.amount),
                'spent': spent,
                'remaining': remaining,
                'utilization': utilization,
                'status': status,
                'start_date': budget.start_date,
                'end_date': budget.end_date,
            })
            
            total_budget_amount += float(budget.amount)
            total_budget_spent += spent
        
        # Overall budget health
        overall_budget_utilization = (total_budget_spent / total_budget_amount * 100) if total_budget_amount > 0 else 0
        
        # Prepare data for budget vs actual chart
        budget_categories = [item['category'] for item in budget_comparison]
        budget_amounts = [item['budget_amount'] for item in budget_comparison]
        actual_amounts = [item['spent'] for item in budget_comparison]

        # ============ ENHANCED KPI CALCULATIONS ============
        
        # Spending Analytics (Last 30 days)
        thirty_days_ago = datetime.now().date() - timedelta(days=30)
        recent_approved_tx = approved_transactions.filter(date__gte=thirty_days_ago)
        recent_spending_30d = abs(recent_approved_tx.filter(amount__lt=0).aggregate(total=Sum('amount'))['total'] or 0)
        
        # Burn rate (daily/weekly/monthly)
        days_with_spending = min(30, (datetime.now().date() - thirty_days_ago).days)
        if days_with_spending > 0:
            daily_burn_rate = recent_spending_30d / days_with_spending
            weekly_burn_rate = daily_burn_rate * 7
            monthly_burn_rate = daily_burn_rate * 30
        else:
            daily_burn_rate = weekly_burn_rate = monthly_burn_rate = 0
        
        # Runway calculation (how many days until funds run out)
        if daily_burn_rate > 0 and remaining_balance > 0:
            runway_days = int(remaining_balance / daily_burn_rate)
            runway_months = runway_days // 30
        else:
            runway_days = 0
            runway_months = 0
        
        # Transaction metrics
        total_tx_count = approved_transactions.count()
        pending_tx_count = Transaction.objects.filter(approved=False).count()
        recent_tx_count = recent_approved_tx.count()
        
        # Average transaction size
        avg_tx_size = (total_spent / total_tx_count) if total_tx_count > 0 else 0
        
        # Spending velocity (transactions per day in last 30 days)
        if days_with_spending > 0:
            velocity = recent_tx_count / days_with_spending
        else:
            velocity = 0
        
        # Category concentration (top category spending percentage)
        if category_spending:
            category_totals = [abs(cat['total']) for cat in category_spending]
            max_category_spend = max(category_totals) if category_totals else 0
            category_concentration = (max_category_spend / total_spent * 100) if total_spent > 0 else 0
        else:
            category_concentration = 0
        
        # Growth rate (compare last 30 days to previous 30 days)
        sixty_days_ago = datetime.now().date() - timedelta(days=60)
        previous_30d_spending = abs(approved_transactions.filter(
            amount__lt=0,
            date__gte=sixty_days_ago,
            date__lt=thirty_days_ago
        ).aggregate(total=Sum('amount'))['total'] or 0)
        
        if previous_30d_spending > 0:
            growth_rate = ((recent_spending_30d - previous_30d_spending) / previous_30d_spending) * 100
        else:
            growth_rate = 0 if recent_spending_30d == 0 else 100
        
        # Income vs Spending Ratio
        if total_income > 0:
            spending_ratio = (total_spent / total_income) * 100
        else:
            spending_ratio = 0
        
        # Category Trend Data - Top 3 categories over last 6 months (for trend visualization)
        last_6_months = datetime.now().date() - timedelta(days=180)
        
        # Get top 3 spending categories
        top_categories_list = list(
            approved_transactions.filter(amount__lt=0, date__gte=last_6_months)
            .values('category')
            .annotate(total=Sum('amount'))
            .order_by('total')[:3]
        )
        
        category_trend_labels = []
        category_trend_datasets = []
        
        if top_categories_list:
            # Build trend data for each top category by month
            for idx, cat_item in enumerate(top_categories_list):
                cat_code = cat_item['category']
                cat_name = category_map.get(cat_code, cat_code)
                
                # Get monthly totals for this category
                monthly_totals = {}
                monthly_txs = approved_transactions.filter(
                    amount__lt=0,
                    category=cat_code,
                    date__gte=last_6_months
                ).order_by('date')
                
                for tx in monthly_txs:
                    month_key = tx.date.strftime('%b %Y')
                    monthly_totals[month_key] = monthly_totals.get(month_key, 0) + abs(tx.amount)
                
                # Only add if has data
                if monthly_totals and idx == 0:
                    category_trend_labels = list(monthly_totals.keys())
                
                if monthly_totals:
                    trend_data = [monthly_totals.get(label, 0) for label in category_trend_labels]
                    colors = ['#F87171', '#60A5FA', '#34D399']
                    category_trend_datasets.append({
                        'label': cat_name,
                        'data': trend_data,
                        'borderColor': colors[idx],
                        'backgroundColor': f'rgba({colors[idx].replace("#", "")}, 0.1)',
                        'tension': 0.3
                    })
        
        # Monthly Comparison Data (This Month vs Last Month)
        today = datetime.now().date()
        first_day_this_month = today.replace(day=1)
        first_day_last_month = (first_day_this_month - timedelta(days=1)).replace(day=1)
        last_day_last_month = first_day_this_month - timedelta(days=1)
        
        this_month_spending = abs(approved_transactions.filter(
            amount__lt=0,
            date__gte=first_day_this_month
        ).aggregate(total=Sum('amount'))['total'] or 0)
        
        last_month_spending = abs(approved_transactions.filter(
            amount__lt=0,
            date__gte=first_day_last_month,
            date__lte=last_day_last_month
        ).aggregate(total=Sum('amount'))['total'] or 0)
        
        month_comparison_data = {
            'this_month': this_month_spending,
            'last_month': last_month_spending,
            'this_month_name': first_day_this_month.strftime('%B'),
            'last_month_name': first_day_last_month.strftime('%B'),
            'monthly_change': ((this_month_spending - last_month_spending) / last_month_spending * 100) if last_month_spending > 0 else 0
        }
        
        # Prepare KPI data for template
        kpis = {
            'burn_rate_daily': daily_burn_rate,
            'burn_rate_weekly': weekly_burn_rate,
            'burn_rate_monthly': monthly_burn_rate,
            'runway_days': runway_days,
            'runway_months': runway_months,
            'total_tx_count': total_tx_count,
            'pending_tx_count': pending_tx_count,
            'recent_tx_count_30d': recent_tx_count,
            'avg_tx_size': avg_tx_size,
            'velocity_per_day': velocity,
            'category_concentration': category_concentration,
            'growth_rate': growth_rate,
            'spending_ratio': spending_ratio,
            'recent_spending_30d': recent_spending_30d,
            'avg_transaction_size': avg_tx_size,
        }

        context = {
            'user': request.user,
            'total_funds': total_funds,
            'total_spent': total_spent,
            'remaining_balance': remaining_balance,
            'transactions': transactions,
            'pending_transactions': pending_transactions,
            'category_spending': category_spending,
            'income_data': income_data,
            'is_treasurer': user_is_treasurer,
            'spending_months': spending_months,
            'spending_amounts': spending_amounts,
            'start_date': start_date_str or '',
            'end_date': end_date_str or '',
            'sponsors_with_tiers': sponsors_with_tiers,
            'management_funds_list': mf_qs.order_by('-date_received'),
            'sponsors_list': sp_qs.order_by('-date_received'),
            # Budget tracking data
            'budget_comparison': budget_comparison,
            'budget_categories': budget_categories,
            'budget_amounts': budget_amounts,
            'actual_amounts': actual_amounts,
            'total_budget_amount': total_budget_amount,
            'total_budget_spent': total_budget_spent,
            'overall_budget_utilization': overall_budget_utilization,
            'budget_exceeded_count': budget_exceeded_count,
            'budget_warning_count': budget_warning_count,
            'has_budgets': budgets.exists(),
            # New KPI data
            'kpis': kpis,
            # Trend and comparison data
            'category_trend_labels': category_trend_labels,
            'category_trend_datasets': category_trend_datasets,
            'month_comparison_data': month_comparison_data,
        }
    except Exception as e:
        context['error'] = f"An unexpected error occurred: {e}"
        logger.error(f"Error in dashboard view: {e}", exc_info=True)

    return render(request, 'tedx_finance/dashboard.html', context)


@login_required
def finance_report(request):
    """
    Printable finance report view with optional date range filters.
    Shows all approved transactions, sponsors, and financial summary.
    """
    try:
        # Parse optional date filters
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        start_date = parse_date(start_date_str)
        end_date = parse_date(end_date_str)

        # Income within range
        mf_qs = ManagementFund.objects.all()
        sp_qs = Sponsor.objects.all()
        if start_date:
            mf_qs = mf_qs.filter(date_received__gte=start_date)
            sp_qs = sp_qs.filter(date_received__gte=start_date)
        if end_date:
            mf_qs = mf_qs.filter(date_received__lte=end_date)
            sp_qs = sp_qs.filter(date_received__lte=end_date)

        total_income = (mf_qs.aggregate(total=Sum('amount'))['total'] or 0) + (sp_qs.aggregate(total=Sum('amount'))['total'] or 0)

        # Sponsor tiers list within selected date range
        sponsors_with_tiers = [
            {
                'name': s.name,
                'amount': float(s.amount),
                'tier': get_sponsor_tier(s.amount)[0]
            }
            for s in sp_qs.order_by('-amount')
        ]

        # Approved transactions within range
        tx_qs = Transaction.objects.filter(approved=True)
        if start_date:
            tx_qs = tx_qs.filter(date__gte=start_date)
        if end_date:
            tx_qs = tx_qs.filter(date__lte=end_date)
        tx_qs = tx_qs.order_by('date')

        total_spent_val = tx_qs.filter(amount__lt=0).aggregate(total=Sum('amount'))['total'] or 0
        total_spent = abs(total_spent_val)
        net_balance = total_income - total_spent

        context = {
            'transactions': tx_qs,
            'total_income': total_income,
            'total_spent': total_spent,
            'net_balance': net_balance,
            'start_date': start_date_str or '',
            'end_date': end_date_str or '',
            'sponsors_with_tiers': sponsors_with_tiers,
        }
    except Exception as e:
        context = {'transactions': [], 'total_income': 0, 'total_spent': 0, 'net_balance': 0, 'error': str(e)}
    return render(request, 'tedx_finance/finance_report.html', context)


@login_required
def export_transactions_pdf(request):
    """
    Export financial report as PDF with comprehensive filtering support.
    
    Features:
    - Supports all transaction filters (search, status, category, date range, amount range, submitted by)
    - Includes sponsors with tier classification (when no specific transaction filters applied)
    - Shows filtered transactions based on user criteria
    - Includes proof images for transactions
    - Comprehensive error handling with user-friendly messages
    
    Args:
        request: HTTP request with optional filter parameters
        
    Returns:
        HttpResponse with PDF attachment or error message
    """
    from xhtml2pdf import pisa
    from django.template.loader import render_to_string
    from django.http import HttpResponse
    import io
    import logging
    import os
    
    logger = logging.getLogger(__name__)
    user_is_treasurer = is_in_group(request.user, 'Treasurer')
    
    # Parse date filters (for sponsor/fund income calculation)
    start_date_str = request.GET.get('start_date', '')
    end_date_str = request.GET.get('end_date', '')
    start_date = parse_date(start_date_str)
    end_date = parse_date(end_date_str)
    
    try:
        # Sponsors with date filtering
        sp_qs = Sponsor.objects.all()
        mf_qs = ManagementFund.objects.all()
        if start_date:
            sp_qs = sp_qs.filter(date_received__gte=start_date)
            mf_qs = mf_qs.filter(date_received__gte=start_date)
        if end_date:
            sp_qs = sp_qs.filter(date_received__lte=end_date)
            mf_qs = mf_qs.filter(date_received__lte=end_date)
        
        sponsor_income = sp_qs.aggregate(total=Sum('amount'))['total'] or 0
        management_income = mf_qs.aggregate(total=Sum('amount'))['total'] or 0
        total_income = sponsor_income + management_income
        
        sponsors_with_tiers = [
            {
                'name': s.name,
                'amount': float(s.amount),
                'tier': get_sponsor_tier(s.amount)[0]
            }
            for s in sp_qs.order_by('-amount')
        ]
        
        # Apply comprehensive transaction filters
        tx_qs = apply_transaction_filters(request, user_is_treasurer=user_is_treasurer)
        tx_qs = tx_qs.order_by('date')
        
        # Calculate totals from filtered transactions
        total_spent_val = tx_qs.filter(amount__lt=0).aggregate(total=Sum('amount'))['total'] or 0
        total_spent = abs(total_spent_val)
        net_balance = total_income - total_spent
        
        # Check if there's data to export
        if not tx_qs.exists() and not sp_qs.exists():
            messages.warning(request, '⚠️ No data found matching your filters.')
            return redirect('tedx_finance:transactions_table')
        
        # Build filter summary for PDF title
        filter_summary = []
        if request.GET.get('search'):
            filter_summary.append(f"Search: {request.GET.get('search')}")
        if request.GET.get('status') and request.GET.get('status') != 'all':
            filter_summary.append(f"Status: {request.GET.get('status').title()}")
        if request.GET.get('category') and request.GET.get('category') != 'all':
            filter_summary.append(f"Category: {request.GET.get('category')}")
        
        context = {
            'transactions': tx_qs,
            'total_income': total_income,
            'total_spent': total_spent,
            'net_balance': net_balance,
            'start_date': start_date_str or '',
            'end_date': end_date_str or '',
            'sponsors_with_tiers': sponsors_with_tiers,
            'filter_summary': ' | '.join(filter_summary) if filter_summary else '',
            'for_pdf': True,
        }
        
    except Exception as e:
        logger.error(f"Error preparing PDF export data: {str(e)}", exc_info=True)
        messages.error(request, f'❌ Error preparing report data: {str(e)}')
        return redirect('tedx_finance:dashboard')
    
    try:
        # Render HTML template
        html_string = render_to_string('tedx_finance/finance_report.html', context, request=request)
        
        # Generate PDF using xhtml2pdf
        result = io.BytesIO()
        pdf = pisa.pisaDocument(io.BytesIO(html_string.encode("UTF-8")), result)
        
        if not pdf.err:
            response = HttpResponse(result.getvalue(), content_type='application/pdf')
            filename = f'tedx_finance_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        else:
            logger.error(f"xhtml2pdf error: {pdf.err}")
            messages.error(request, '❌ Error generating PDF. Please check the report template.')
            return redirect('tedx_finance:dashboard')
            
    except Exception as e:
        logger.error(f"Error generating PDF: {str(e)}", exc_info=True)
        messages.error(request, f'❌ Failed to generate PDF: {str(e)}')
        return redirect('tedx_finance:dashboard')


# --- Transaction Management ---
@login_required
def add_transaction(request):
    if request.method == 'POST':
        form = TransactionForm(request.user, request.POST, request.FILES)
        if form.is_valid():
            transaction = form.save(commit=False)
            transaction.created_by = request.user
            transaction.save()
            messages.success(request, 'Transaction submitted for approval!')
            return redirect('tedx_finance:dashboard')
        else:
            messages.error(request, 'Please correct the errors below and resubmit the transaction.')
    else:
        form = TransactionForm(request.user)
    context = {
        'form': form,
        'form_title': 'Add a New Transaction',
        'form_button': 'Save Transaction',
        'is_treasurer': is_in_group(request.user, 'Treasurer'),
    }
    return render(request, 'tedx_finance/add_transaction.html', context)

@permission_required('tedx_finance.change_transaction', raise_exception=True)
def approve_transaction(request, pk):
    from .utils import log_audit_action, create_notification
    
    transaction = get_object_or_404(Transaction, pk=pk)
    if request.method == 'POST':
        transaction.approved = True
        transaction.save()
        
        # Log audit action
        log_audit_action(
            request.user,
            'approve_transaction',
            'Transaction',
            transaction.id,
            f"Approved transaction: {transaction.title} (₹{transaction.amount})",
            getattr(request, 'client_ip', None)
        )
        
        # Create notification for transaction creator if different from approver
        if transaction.created_by and transaction.created_by != request.user:
            create_notification(
                transaction.created_by,
                'transaction_approved',
                'Transaction Approved',
                f"Your transaction '{transaction.title}' has been approved by {request.user.username}.",
                'Transaction',
                transaction.id
            )
        
        messages.success(request, f"Transaction '{transaction.title}' approved.")
    return redirect('tedx_finance:dashboard')

@permission_required('tedx_finance.change_transaction', raise_exception=True)
def edit_transaction(request, pk):
    """Edit an existing transaction (treasurers only, audit trail via history)."""
    tx = get_object_or_404(Transaction, pk=pk)
    if request.method == 'POST':
        form = TransactionForm(request.user, request.POST, request.FILES, instance=tx)
        if form.is_valid():
            form.save()
            messages.success(request, 'Transaction updated successfully.')
            return redirect('tedx_finance:transactions_table')
        else:
            messages.error(request, 'Please correct the errors below and resubmit the transaction.')
    else:
        form = TransactionForm(request.user, instance=tx)
    return render(request, 'tedx_finance/add_transaction.html', {
        'form': form,
        'form_title': f"Edit Transaction: {tx.title}",
        'form_button': 'Update Transaction'
    })

@permission_required('tedx_finance.delete_transaction', raise_exception=True)
def reject_transaction(request, pk):
    from .utils import log_audit_action, create_notification
    
    transaction = get_object_or_404(Transaction, pk=pk)
    if request.method == 'POST':
        title = transaction.title
        amount = transaction.amount
        created_by = transaction.created_by
        
        # Log audit action before deletion
        log_audit_action(
            request.user,
            'reject_transaction',
            'Transaction',
            transaction.id,
            f"Rejected and deleted transaction: {title} (₹{amount})",
            getattr(request, 'client_ip', None)
        )
        
        # Create notification for transaction creator
        if created_by:
            create_notification(
                created_by,
                'transaction_rejected',
                'Transaction Rejected',
                f"Your transaction '{title}' has been rejected and removed by {request.user.username}.",
                'Transaction',
                pk
            )
        
        transaction.delete()
        messages.warning(request, f"Transaction '{title}' has been rejected and deleted.")
    return redirect('tedx_finance:dashboard')

@permission_required('tedx_finance.change_transaction', raise_exception=True)
def bulk_approve_transactions(request):
    """Bulk approve multiple transactions."""
    if request.method == 'POST':
        import json
        try:
            data = json.loads(request.body)
            transaction_ids = data.get('ids', [])
            
            if not transaction_ids:
                return JsonResponse({'success': False, 'error': 'No transaction IDs provided'}, status=400)
            
            # Approve all transactions
            transactions = Transaction.objects.filter(pk__in=transaction_ids)
            count = transactions.update(approved=True)
            
            return JsonResponse({
                'success': True,
                'count': count,
                'message': f'{count} transaction(s) approved successfully'
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=405)

@permission_required('tedx_finance.delete_transaction', raise_exception=True)
def bulk_reject_transactions(request):
    """Bulk reject (delete) multiple transactions."""
    if request.method == 'POST':
        import json
        try:
            data = json.loads(request.body)
            transaction_ids = data.get('ids', [])
            
            if not transaction_ids:
                return JsonResponse({'success': False, 'error': 'No transaction IDs provided'}, status=400)
            
            # Delete all transactions
            transactions = Transaction.objects.filter(pk__in=transaction_ids)
            count = transactions.count()
            transactions.delete()
            
            return JsonResponse({
                'success': True,
                'count': count,
                'message': f'{count} transaction(s) rejected and deleted successfully'
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=405)

# --- Income Management ---
@permission_required('tedx_finance.add_managementfund', raise_exception=True)
def add_management_fund(request):
    from .utils import log_audit_action
    
    if request.method == 'POST':
        form = ManagementFundForm(request.POST)
        if form.is_valid():
            fund = form.save()
            
            # Log audit action
            log_audit_action(
                request.user,
                'create_fund',
                'ManagementFund',
                fund.id,
                f"Created management fund: ₹{fund.amount} on {fund.date_received}",
                getattr(request, 'client_ip', None)
            )
            
            messages.success(request, 'Management fund updated!')
            return redirect('tedx_finance:dashboard')
        else:
            messages.error(request, 'Please correct the errors below and resubmit the fund details.')
    else:
        form = ManagementFundForm()
    context = {'form': form, 'form_title': 'Add Management Fund', 'form_button': 'Save Fund'}
    return render(request, 'tedx_finance/add_management_fund.html', context)

@permission_required('tedx_finance.change_managementfund', raise_exception=True)
def edit_management_fund(request, pk):
    """Edit an existing management fund (treasurers only)."""
    from .utils import log_audit_action
    
    fund = get_object_or_404(ManagementFund, pk=pk)
    if request.method == 'POST':
        form = ManagementFundForm(request.POST, instance=fund)
        if form.is_valid():
            old_amount = fund.amount
            form.save()
            
            # Log audit action
            log_audit_action(
                request.user,
                'update_fund',
                'ManagementFund',
                fund.id,
                f"Updated management fund from ₹{old_amount} to ₹{fund.amount}",
                getattr(request, 'client_ip', None)
            )
            
            messages.success(request, 'Management fund updated successfully.')
            return redirect('tedx_finance:dashboard')
        else:
            messages.error(request, 'Please correct the errors below and resubmit the fund details.')
    else:
        form = ManagementFundForm(instance=fund)
    return render(request, 'tedx_finance/add_management_fund.html', {
        'form': form,
        'form_title': f"Edit Management Fund: ₹{fund.amount}",
        'form_button': 'Update Fund'
    })

@permission_required('tedx_finance.delete_managementfund', raise_exception=True)
def delete_management_fund(request, pk):
    """Delete a management fund entry (treasurers only)."""
    from .utils import log_audit_action
    
    fund = get_object_or_404(ManagementFund, pk=pk)
    if request.method == 'POST':
        amount = fund.amount
        date = fund.date_received
        
        # Log audit action before deletion
        log_audit_action(
            request.user,
            'delete_fund',
            'ManagementFund',
            fund.id,
            f"Deleted management fund: ₹{amount} from {date}",
            getattr(request, 'client_ip', None)
        )
        
        fund.delete()
        messages.warning(request, f'Management fund of ₹{amount} has been deleted.')
    return redirect('tedx_finance:dashboard')

@permission_required('tedx_finance.add_sponsor', raise_exception=True)
def add_sponsor(request):
    if request.method == 'POST':
        form = SponsorForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'New sponsor added successfully!')
            return redirect('tedx_finance:dashboard')
        else:
            messages.error(request, 'Please correct the errors below and resubmit the sponsor details.')
    else:
        form = SponsorForm()
    context = {'form': form, 'form_title': 'Add New Sponsor', 'form_button': 'Save Sponsor'}
    return render(request, 'tedx_finance/add_sponsor.html', context)

@permission_required('tedx_finance.change_sponsor', raise_exception=True)
def edit_sponsor(request, pk):
    """Edit an existing sponsor (treasurers only)."""
    sponsor = get_object_or_404(Sponsor, pk=pk)
    if request.method == 'POST':
        form = SponsorForm(request.POST, request.FILES, instance=sponsor)
        if form.is_valid():
            form.save()
            messages.success(request, 'Sponsor updated successfully.')
            return redirect('tedx_finance:dashboard')
        else:
            messages.error(request, 'Please correct the errors below and resubmit the sponsor details.')
    else:
        form = SponsorForm(instance=sponsor)
    return render(request, 'tedx_finance/add_sponsor.html', {
        'form': form,
        'form_title': f"Edit Sponsor: {sponsor.name}",
        'form_button': 'Update Sponsor'
    })

@permission_required('tedx_finance.delete_sponsor', raise_exception=True)
def delete_sponsor(request, pk):
    """Delete a sponsor entry (treasurers only)."""
    from .utils import log_audit_action
    
    sponsor = get_object_or_404(Sponsor, pk=pk)
    if request.method == 'POST':
        name = sponsor.name
        amount = sponsor.amount
        
        # Log audit action before deletion
        log_audit_action(
            request.user,
            'delete_fund',
            'Sponsor',
            sponsor.id,
            f"Deleted sponsor: {name} (₹{amount})",
            getattr(request, 'client_ip', None)
        )
        
        sponsor.delete()
        messages.warning(request, f'Sponsor "{name}" has been deleted.')
    return redirect('tedx_finance:dashboard')

# --- Exports ---
@login_required
def export_transactions_xlsx(request):
    """
    Export transactions to Excel format with comprehensive filtering support.
    
    Features:
    - Supports all transaction filters (search, status, category, date range, amount range, submitted by)
    - Formatted Excel with headers and proper column widths
    - Includes transaction metadata (date, title, category, amount, submitter, status)
    - Summary row with totals and transaction count
    - Comprehensive error handling
    
    Args:
        request: HTTP request with optional filter parameters
        
    Returns:
        HttpResponse with Excel attachment or redirect with error message
    """
    import logging
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    
    logger = logging.getLogger(__name__)
    user_is_treasurer = is_in_group(request.user, 'Treasurer')
    
    try:
        # Apply all filters using helper function
        transactions = apply_transaction_filters(request, user_is_treasurer=user_is_treasurer)
        transactions = transactions.order_by('date')
        
        # Check if there's data to export
        if not transactions.exists():
            messages.warning(request, '⚠️ No transactions found matching your filters.')
            return redirect('tedx_finance:transactions_table')
        
        # Create Excel workbook
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        
        # Build filename with filter summary
        filter_parts = []
        if request.GET.get('search'):
            filter_parts.append('filtered')
        if request.GET.get('status') and request.GET.get('status') != 'all':
            filter_parts.append(request.GET.get('status'))
        if request.GET.get('category') and request.GET.get('category') != 'all':
            filter_parts.append(request.GET.get('category'))
        
        filter_suffix = f"_{'_'.join(filter_parts)}" if filter_parts else ''
        filename = f'tedx_transactions{filter_suffix}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Transactions'
        
        # Header row with styling
        columns = ['Date', 'Title', 'Category', 'Amount', 'Status', 'Submitted By']
        worksheet.append(columns)
        
        # Style header row
        header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data rows
        for tx in transactions:
            row = [
                tx.date,
                tx.title,
                tx.get_category_display(),
                float(tx.amount),
                'Approved' if tx.approved else 'Pending',
                tx.created_by.username if tx.created_by else 'N/A'
            ]
            worksheet.append(row)
        
        # Auto-adjust column widths
        for col_num, column_cells in enumerate(worksheet.columns, 1):
            max_length = 0
            column = get_column_letter(col_num)
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column].width = adjusted_width
        
        # Add summary row
        row_count = worksheet.max_row
        total_amount = sum(float(tx.amount) for tx in transactions)
        summary_row = [
            '',
            '',
            'TOTAL',
            total_amount,
            '',
            f'{transactions.count()} transactions'
        ]
        worksheet.append(summary_row)
        
        # Style summary row
        summary_fill = PatternFill(start_color='E5E7EB', end_color='E5E7EB', fill_type='solid')
        summary_font = Font(bold=True)
        for col_num in range(1, 6):
            cell = worksheet.cell(row=row_count + 1, column=col_num)
            cell.fill = summary_fill
            cell.font = summary_font
        
        workbook.save(response)
        return response
        
    except Exception as e:
        logger.error(f"Error exporting transactions to Excel: {str(e)}", exc_info=True)
        messages.error(request, f'❌ Failed to export Excel file: {str(e)}')
        return redirect('tedx_finance:dashboard')


@login_required
def export_transactions_with_proofs(request):
    """
    Export approved transactions as Excel + all proof images in a ZIP file.
    
    Features:
    - Includes Excel report with transaction data
    - Bundles all proof images organized by transaction
    - Optional date range filtering
    - Creates organized folder structure: proofs/TransactionTitle_Date/
    
    Args:
        request: HTTP request with optional start_date and end_date GET parameters
        
    Returns:
        HttpResponse with ZIP attachment containing Excel + proof images
    """
    import logging
    import zipfile
    import io
    import os
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    from django.core.files.storage import default_storage
    
    logger = logging.getLogger(__name__)
    
    try:
        # Parse optional filters
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        start_date = parse_date(start_date_str)
        end_date = parse_date(end_date_str)

        # Query approved transactions
        transactions = Transaction.objects.filter(approved=True)
        if start_date:
            transactions = transactions.filter(date__gte=start_date)
        if end_date:
            transactions = transactions.filter(date__lte=end_date)
        transactions = transactions.order_by('date')
        
        if not transactions.exists():
            messages.warning(request, '⚠️ No approved transactions found for the selected date range.')
            return redirect('tedx_finance:transactions_table')
        
        # Create in-memory ZIP file
        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # --- 1. Create Excel report ---
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = 'Transactions'
            
            # Header row with styling
            columns = ['Date', 'Title', 'Category', 'Amount', 'Submitted By', 'Proof File']
            worksheet.append(columns)
            
            # Style header row
            header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=12)
            for col_num, column_title in enumerate(columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Data rows with proof file references
            for tx in transactions:
                proof_filename = ''
                if tx.proof:
                    proof_filename = os.path.basename(tx.proof.name)
                
                row = [
                    tx.date,
                    tx.title,
                    tx.get_category_display(),
                    float(tx.amount),
                    tx.created_by.username if tx.created_by else 'N/A',
                    proof_filename or 'No proof uploaded'
                ]
                worksheet.append(row)
            
            # Auto-adjust column widths
            for col_num, column_cells in enumerate(worksheet.columns, 1):
                max_length = 0
                column = get_column_letter(col_num)
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column].width = adjusted_width
            
            # Add summary row
            row_count = worksheet.max_row
            total_amount = sum(float(tx.amount) for tx in transactions)
            summary_row = [
                '',
                '',
                'TOTAL',
                total_amount,
                f'{transactions.count()} transactions',
                ''
            ]
            worksheet.append(summary_row)
            
            # Style summary row
            summary_fill = PatternFill(start_color='E5E7EB', end_color='E5E7EB', fill_type='solid')
            summary_font = Font(bold=True)
            for col_num in range(1, 7):
                cell = worksheet.cell(row=row_count + 1, column=col_num)
                cell.fill = summary_fill
                cell.font = summary_font
            
            # Save Excel to buffer
            excel_buffer = io.BytesIO()
            workbook.save(excel_buffer)
            excel_buffer.seek(0)
            
            # Add Excel to ZIP
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            zip_file.writestr(f'tedx_transactions_{timestamp}.xlsx', excel_buffer.read())
            
            # --- 2. Add proof images to ZIP ---
            proofs_added = 0
            for tx in transactions:
                if tx.proof:
                    try:
                        # Create organized folder structure
                        safe_title = "".join(c for c in tx.title if c.isalnum() or c in (' ', '-', '_')).strip()
                        folder_name = f"proofs/{safe_title}_{tx.date}"
                        file_extension = os.path.splitext(tx.proof.name)[1]
                        proof_path = f"{folder_name}/proof{file_extension}"
                        
                        # Read proof file and add to ZIP
                        if default_storage.exists(tx.proof.name):
                            with default_storage.open(tx.proof.name, 'rb') as proof_file:
                                zip_file.writestr(proof_path, proof_file.read())
                                proofs_added += 1
                    except Exception as e:
                        logger.warning(f"Could not add proof for transaction {tx.id}: {str(e)}")
            
            # Add a README file
            readme_content = f"""TEDx Finance Report with Proofs
=====================================

Export Date: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
Date Range: {start_date or 'All'} to {end_date or 'All'}

Contents:
---------
1. tedx_transactions_{timestamp}.xlsx - Full transaction report
2. proofs/ folder - {proofs_added} proof images organized by transaction

File Structure:
---------------
proofs/
  ├── TransactionTitle_Date/
  │   └── proof.jpg (or .pdf, .png, etc.)
  └── ...

Notes:
------
- Only approved transactions are included
- Proof images are organized by transaction title and date
- If a transaction has no proof, it won't appear in the proofs folder
- All amounts are in Indian Rupees (₹)

For questions, contact the TEDx Finance Team.
"""
            zip_file.writestr('README.txt', readme_content)
        
        # Prepare ZIP response
        zip_buffer.seek(0)
        response = HttpResponse(zip_buffer.read(), content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename="tedx_report_with_proofs_{timestamp}.zip"'
        
        messages.success(request, f'✅ Exported {transactions.count()} transactions with {proofs_added} proof files!')
        return response
        
    except Exception as e:
        logger.error(f"Error exporting ZIP with proofs: {str(e)}", exc_info=True)
        messages.error(request, f'❌ Failed to export ZIP file: {str(e)}')
        return redirect('tedx_finance:dashboard')


@permission_required('tedx_finance.add_transaction', raise_exception=True)
def import_transactions(request):
    """
    Bulk import transactions from .xlsx file.
    Validates file type, size (max 5MB), and data format.
    Expected columns: Date, Title, Category, Amount, SubmittedBy (optional)
    """
    report = None
    MAX_FILE_SIZE_MB = 5  # Maximum file size in megabytes
    
    if request.method == 'POST':
        uploaded_file = request.FILES.get('file')
        
        # Validate file type
        if not uploaded_file or not uploaded_file.name.endswith('.xlsx'):
            messages.error(request, 'Please upload a valid .xlsx file.')
        # Validate file size
        elif uploaded_file.size > MAX_FILE_SIZE_MB * 1024 * 1024:
            messages.error(request, f'File size exceeds {MAX_FILE_SIZE_MB}MB limit.')
        else:
            try:
                wb = openpyxl.load_workbook(uploaded_file)
                ws = wb.active
                errors = []
                created = 0
                from django.contrib.auth.models import User
                
                # Process each row (skip header row)
                for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if not row[0]:  # Skip empty rows
                        continue
                    try:
                        # Parse row data
                        date_val = row[0] if isinstance(row[0], str) else row[0].strftime('%Y-%m-%d') if hasattr(row[0], 'strftime') else str(row[0])
                        title = str(row[1]) if len(row) > 1 and row[1] else ''
                        category = str(row[2]) if len(row) > 2 and row[2] else ''
                        amount = float(row[3]) if len(row) > 3 and row[3] else 0.0
                        username = str(row[4]).strip() if len(row) > 4 and row[4] else None
                        
                        # Validate required fields
                        if not title or not category:
                            errors.append({'row': idx, 'message': 'Missing title or category'})
                            continue
                        
                        # Ensure category exists in Category table (create if missing)
                        if category:
                            try:
                                Category.objects.get_or_create(name=category)
                            except Exception:
                                pass
                        
                        # Find or default user
                        user = None
                        if username:
                            try:
                                user = User.objects.get(username=username)
                            except User.DoesNotExist:
                                pass
                        if not user:
                            user = request.user
                        
                        # Create transaction
                        Transaction.objects.create(
                            title=title,
                            amount=amount,
                            category=category,
                            date=date_val,
                            created_by=user,
                            approved=False
                        )
                        created += 1
                    except Exception as e:
                        errors.append({'row': idx, 'message': str(e)})
                
                report = {'created': created, 'errors': errors}
                if created > 0:
                    messages.success(request, f'{created} transactions imported successfully.')
                if errors:
                    messages.warning(request, f'{len(errors)} rows had errors.')
            except Exception as e:
                messages.error(request, f'Error reading file: {e}')
    
    try:
        dynamic = list(Category.objects.all().values_list('name', 'name'))
    except Exception:
        dynamic = []
    default_choices = list(getattr(Transaction, 'CATEGORY_CHOICES', []))
    seen = set()
    categories = []
    for val, label in dynamic + default_choices:
        if val not in seen:
            categories.append((val, label))
            seen.add(val)

    context = {
        'categories': categories,
        'report': report,
    }
    return render(request, 'tedx_finance/import_transactions.html', context)


@login_required
def manage_categories(request):
    """Simple manager to add/remove custom categories (Treasurer only)."""
    user_is_treasurer = is_in_group(request.user, 'Treasurer')
    if not user_is_treasurer:
        messages.error(request, 'Only treasurers can manage categories.')
        return redirect('tedx_finance:dashboard')

    # Add new category
    if request.method == 'POST':
        name = (request.POST.get('name') or '').strip()
        if not name:
            messages.error(request, 'Category name cannot be empty.')
        elif len(name) > 50:
            messages.error(request, 'Category name must be 50 characters or fewer.')
        else:
            obj, created = Category.objects.get_or_create(name=name)
            if created:
                messages.success(request, f'Category "{name}" added.')
            else:
                messages.info(request, f'Category "{name}" already exists.')
        return redirect('tedx_finance:manage_categories')

    # Delete category
    delete_id = request.GET.get('delete')
    if delete_id:
        try:
            cat = Category.objects.get(id=delete_id)
            cat.delete()
            invalidate_category_cache()
            messages.warning(request, f'Category "{cat.name}" deleted.')
            return redirect('tedx_finance:manage_categories')
        except Category.DoesNotExist:
            messages.error(request, 'Category not found.')

    categories = Category.objects.all()
    return render(request, 'tedx_finance/manage_categories.html', {
        'categories': categories,
        'is_treasurer': user_is_treasurer,
    })


@login_required
def proof_gallery(request):
    """
    Gallery view of all transaction proofs with thumbnails and lightbox.
    Shows only approved transactions with uploaded proof files.
    Includes filtering by category and date range.
    """
    user_is_treasurer = is_in_group(request.user, 'Treasurer')
    
    # Parse optional filters
    search_query = request.GET.get('search', '').strip()
    category_filter = request.GET.get('category', '')
    start_date_str = request.GET.get('start_date', '')
    end_date_str = request.GET.get('end_date', '')
    tx_id_param = request.GET.get('tx_id', '')  # For auto-opening specific transaction
    start_date = parse_date(start_date_str)
    end_date = parse_date(end_date_str)
    
    # Query transactions with proofs
    transactions = Transaction.objects.filter(approved=True, proof__isnull=False).exclude(proof='')
    transactions = transactions.select_related('created_by')
    
    if search_query:
        transactions = transactions.filter(
            Q(title__icontains=search_query) | 
            Q(created_by__username__icontains=search_query)
        )
    if category_filter:
        transactions = transactions.filter(category=category_filter)
    if start_date:
        transactions = transactions.filter(date__gte=start_date)
    if end_date:
        transactions = transactions.filter(date__lte=end_date)
    
    transactions = transactions.order_by('-date')
    
    # Get unique categories for filter dropdown
    categories = get_cached_category_choices()
    
    context = {
        'transactions': transactions,
        'categories': categories,
        'search_query': search_query,
        'category_filter': category_filter,
        'start_date': start_date_str,
        'end_date': end_date_str,
        'tx_id_param': tx_id_param,
        'is_treasurer': user_is_treasurer,
    }
    return render(request, 'tedx_finance/proof_gallery.html', context)


@login_required
@permission_required('tedx_finance.change_transaction', raise_exception=True)
def bulk_upload_proofs(request):
    """
    Bulk upload proofs for multiple transactions.
    Allows treasurers to upload multiple proof files at once.
    """
    if request.method == 'POST':
        files = request.FILES.getlist('proof_files')
        transaction_ids = request.POST.getlist('transaction_ids')
        
        if not files:
            messages.error(request, 'No files uploaded.')
            return redirect('tedx_finance:proof_gallery')
        
        success_count = 0
        error_count = 0
        error_details = []
        
        # Handle mapping of files to transactions
        for i, file in enumerate(files):
            try:
                # Validate each file before attaching
                validate_file_size(file)
                validate_file_extension(file)
            except ValidationError as ve:
                error_count += 1
                detail = ve.messages[0] if hasattr(ve, 'messages') and ve.messages else 'Invalid file'
                error_details.append(f"{file.name}: {detail}")
                continue
            try:
                # If transaction IDs provided, map files to transactions
                if transaction_ids and i < len(transaction_ids):
                    tx_id = transaction_ids[i]
                    if not tx_id:
                        raise Transaction.DoesNotExist
                    tx = Transaction.objects.get(id=tx_id, approved=False)
                    tx.proof = file
                    tx.save()
                    success_count += 1
                    
                    # Create notification
                    from .utils import create_notification
                    if tx.created_by:
                        create_notification(
                            tx.created_by,
                            'transaction_approved',
                            'Proof Uploaded',
                            f"Proof has been uploaded for your transaction '{tx.title}'.",
                            'Transaction',
                            tx.id
                        )
                else:
                    # Auto-match by filename pattern (e.g., "transaction_123_proof.pdf")
                    import re
                    match = re.search(r'transaction[_-](\d+)', file.name, re.IGNORECASE)
                    if match:
                        tx_id = match.group(1)
                        tx = Transaction.objects.get(id=tx_id)
                        tx.proof = file
                        tx.save()
                        success_count += 1
                    else:
                        error_count += 1
                        logger.warning(f"Could not match file {file.name} to any transaction")
                        
            except Transaction.DoesNotExist:
                error_count += 1
                error_details.append(f"{file.name}: Transaction not found or not pending")
                logger.error(f"Transaction not found for file {file.name}")
            except Exception as e:
                error_count += 1
                error_details.append(f"{file.name}: {str(e)}")
                logger.error(f"Error uploading proof: {str(e)}")
        
        if success_count > 0:
            messages.success(request, f'Successfully uploaded {success_count} proof(s).')
        if error_count > 0:
            preview = '; '.join(error_details[:3]) if error_details else ''
            more = f" (+{error_count - len(error_details[:3])} more)" if error_count > len(error_details[:3]) else ''
            messages.warning(request, f'{error_count} file(s) could not be processed. {preview}{more}')
        
        return redirect('tedx_finance:proof_gallery')
    
    # GET: Show upload form with pending transactions
        pending_transactions = Transaction.objects.filter(
            approved=False
        ).select_related('created_by').order_by('-date')[:50]  # Limit to recent 50
    
    context = {
        'pending_transactions': pending_transactions,
    }
    return render(request, 'tedx_finance/bulk_upload_proofs.html', context)


@login_required
def quick_add_category(request):
    """AJAX: quick add a new Category (Treasurer only)."""
    if not is_in_group(request.user, 'Treasurer'):
        return JsonResponse({'success': False, 'error': 'Forbidden'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid method'}, status=405)
    import json
    try:
        data = json.loads(request.body or '{}')
        name = (data.get('name') or '').strip()
        if not name:
            return JsonResponse({'success': False, 'error': 'Name is required'}, status=400)
        if len(name) > 50:
            return JsonResponse({'success': False, 'error': 'Name must be <= 50 chars'}, status=400)
        cat, created = Category.objects.get_or_create(name=name)
        invalidate_category_cache()
        # Build merged categories list (dynamic + defaults) preserving order
        try:
            dynamic = list(Category.objects.all().values_list('name', 'name'))
        except Exception:
            dynamic = []
        default_choices = list(getattr(Transaction, 'CATEGORY_CHOICES', []))
        seen = set()
        merged = []
        for val, label in dynamic + default_choices:
            if val not in seen:
                merged.append({'value': val, 'label': label})
                seen.add(val)
        return JsonResponse({'success': True, 'id': cat.id, 'name': cat.name, 'categories': merged})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def quick_rename_category(request):
    """AJAX: rename a Category and cascade to Transaction.category strings (Treasurer only)."""
    if not is_in_group(request.user, 'Treasurer'):
        return JsonResponse({'success': False, 'error': 'Forbidden'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid method'}, status=405)
    import json
    try:
        data = json.loads(request.body or '{}')
        cat_id = data.get('id')
        old_name = (data.get('old_name') or '').strip()
        new_name = (data.get('new_name') or '').strip()
        if not new_name:
            return JsonResponse({'success': False, 'error': 'New name is required'}, status=400)
        if len(new_name) > 50:
            return JsonResponse({'success': False, 'error': 'Name must be <= 50 chars'}, status=400)
        # Check conflict
        if Category.objects.filter(name=new_name).exclude(id=cat_id).exists():
            return JsonResponse({'success': False, 'error': 'Category with this name already exists'}, status=409)
        cat = None
        if cat_id:
            cat = Category.objects.filter(id=cat_id).first()
        if not cat and old_name:
            cat, _ = Category.objects.get_or_create(name=old_name)
        if not cat:
            return JsonResponse({'success': False, 'error': 'Category not found'}, status=404)
        # Rename
        prev_name = cat.name
        cat.name = new_name
        cat.save()
        invalidate_category_cache()
        # Update Transaction rows that used the old string value
        Transaction.objects.filter(category=prev_name).update(category=new_name)
        # Build merged list
        try:
            dynamic = list(Category.objects.all().values_list('name', 'name'))
        except Exception:
            dynamic = []
        default_choices = list(getattr(Transaction, 'CATEGORY_CHOICES', []))
        seen = set()
        merged = []
        for val, label in dynamic + default_choices:
            if val not in seen:
                merged.append({'value': val, 'label': label})
                seen.add(val)
        return JsonResponse({'success': True, 'id': cat.id, 'name': cat.name, 'categories': merged})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def export_proofs_to_csv(request):
    """
    Export proof gallery data to CSV file.
    Includes transaction details and proof file URLs.
    """
    import csv
    
    # Apply same filters as proof_gallery view
    category_filter = request.GET.get('category', '')
    start_date_str = request.GET.get('start_date', '')
    end_date_str = request.GET.get('end_date', '')
    start_date = parse_date(start_date_str)
    end_date = parse_date(end_date_str)
    
    transactions = Transaction.objects.filter(approved=True, proof__isnull=False).exclude(proof='')
    
    if category_filter:
        transactions = transactions.filter(category=category_filter)
    if start_date:
        transactions = transactions.filter(date__gte=start_date)
    if end_date:
        transactions = transactions.filter(date__lte=end_date)
    
    transactions = transactions.order_by('-date')
    
    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="proof_gallery_{timestamp}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Date', 'Title', 'Category', 'Amount (₹)', 'Description', 'Proof File'])
    
    for tx in transactions:
        proof_url = request.build_absolute_uri(tx.proof.url) if tx.proof else ''
        writer.writerow([
            tx.date.strftime('%Y-%m-%d'),
            tx.title,
            tx.category,
            f"{tx.amount:.2f}",
            tx.description or '',
            proof_url
        ])
    
    return response


@login_required  
def export_proofs_to_pdf(request):
    """
    Export proof gallery data to PDF file.
    Creates a formatted PDF report with transaction details.
    """
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER
    except ImportError:
        messages.error(request, 'PDF export requires reportlab library. Please install: pip install reportlab')
        return redirect('proof_gallery')
    
    # Apply same filters as proof_gallery view
    category_filter = request.GET.get('category', '')
    start_date_str = request.GET.get('start_date', '')
    end_date_str = request.GET.get('end_date', '')
    start_date = parse_date(start_date_str)
    end_date = parse_date(end_date_str)
    
    transactions = Transaction.objects.filter(approved=True, proof__isnull=False).exclude(proof='')
    
    if category_filter:
        transactions = transactions.filter(category=category_filter)
    if start_date:
        transactions = transactions.filter(date__gte=start_date)
    if end_date:
        transactions = transactions.filter(date__lte=end_date)
    
    transactions = transactions.order_by('-date')
    
    # Create PDF response
    response = HttpResponse(content_type='application/pdf')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="proof_gallery_{timestamp}.pdf"'
    
    # Create PDF document
    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#DC2626'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    elements.append(Paragraph('TEDx Proof Gallery Report', title_style))
    elements.append(Spacer(1, 0.3*inch))
    
    # Filters info
    if category_filter or start_date or end_date:
        info_style = styles['Normal']
        filter_text = 'Filters: '
        filters_applied = []
        if category_filter:
            filters_applied.append(f"Category: {category_filter}")
        if start_date:
            filters_applied.append(f"From: {start_date.strftime('%Y-%m-%d')}")
        if end_date:
            filters_applied.append(f"To: {end_date.strftime('%Y-%m-%d')}")
        filter_text += ', '.join(filters_applied)
        elements.append(Paragraph(filter_text, info_style))
        elements.append(Spacer(1, 0.2*inch))
    
    # Table data
    data = [['Date', 'Title', 'Category', 'Amount (₹)']]
    total_amount = 0
    
    for tx in transactions:
        data.append([
            tx.date.strftime('%Y-%m-%d'),
            tx.title[:30] + '...' if len(tx.title) > 30 else tx.title,
            tx.category[:20] if tx.category else '',
            f"₹{tx.amount:,.2f}"
        ])
        total_amount += tx.amount
    
    # Add total row
    data.append(['', '', 'Total:', f"₹{total_amount:,.2f}"])
    
    # Create table
    table = Table(data, colWidths=[1.5*inch, 2.5*inch, 1.8*inch, 1.5*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#DC2626')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('GRID', (0, 0), (-1, -2), 1, colors.black),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FEE2E2')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('LINEABOVE', (0, -1), (-1, -1), 2, colors.HexColor('#DC2626')),
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Footer
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey
    )
    elements.append(Paragraph(f'Generated on {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}', footer_style))
    elements.append(Paragraph(f'Total Transactions: {len(transactions)}', footer_style))
    
    # Build PDF
    doc.build(elements)
    return response


# ============================================================================
# SETTINGS
# ============================================================================


@login_required
def settings_view(request):
    """User settings for theme and notification preferences."""
    prefs, _ = UserPreference.objects.get_or_create(user=request.user)
    if request.method == 'POST':
        form = UserPreferenceForm(request.POST, instance=prefs)
        if form.is_valid():
            prefs = form.save()
            # Persist theme for subsequent requests and JS bootstrap
            request.session['theme_pref'] = prefs.theme
            response = redirect('tedx_finance:settings')
            response.set_cookie('theme', prefs.theme, max_age=60 * 60 * 24 * 365, samesite='Lax')
            messages.success(request, 'Settings saved successfully.')
            return response
        messages.error(request, 'Please fix the errors below and save again.')
    else:
        form = UserPreferenceForm(instance=prefs)

    context = {
        'form': form,
        'preferences': prefs,
        'current_theme': prefs.theme,
    }
    return render(request, 'tedx_finance/settings.html', context)


# ============================================================================
# NOTIFICATIONS VIEWS
# ============================================================================

@login_required
def notifications_list(request):
    """Display all notifications for the logged-in user with pagination."""
    from .models_improvements import Notification
    
    notifications = Notification.objects.filter(user=request.user).select_related('user')
    unread_count = notifications.filter(is_read=False).count()
    
    # Pagination
    paginator = Paginator(notifications, 20)
    page_number = request.GET.get('page', 1)
    try:
        page_obj = paginator.get_page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.get_page(1)
    except EmptyPage:
        page_obj = paginator.get_page(paginator.num_pages)
    
    context = {
        'page_obj': page_obj,
        'notifications': page_obj.object_list,
        'unread_count': unread_count,
        'total_notifications': notifications.count(),
    }
    
    return render(request, 'tedx_finance/notifications.html', context)


@login_required
def get_unread_notifications_count(request):
    """API endpoint to get unread notifications count (for real-time updates)."""
    from django.http import JsonResponse
    from .models_improvements import Notification
    
    unread_count = Notification.objects.filter(user=request.user, is_read=False).count()
    return JsonResponse({'unread_count': unread_count})


@login_required
def mark_notification_read(request, pk):
    """Mark a single notification as read."""
    from django.http import JsonResponse
    from .models_improvements import Notification
    
    try:
        notification = Notification.objects.get(pk=pk, user=request.user)
        notification.is_read = True
        notification.save()
        
        # Get remaining unread count
        unread_count = Notification.objects.filter(user=request.user, is_read=False).count()
        
        return JsonResponse({
            'success': True,
            'unread_count': unread_count,
            'message': 'Notification marked as read'
        })
    except Notification.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Notification not found'}, status=404)


@login_required
def mark_all_notifications_read(request):
    """Mark all notifications as read for the user."""
    from django.http import JsonResponse
    from .models_improvements import Notification
    
    if request.method == 'POST':
        unread_notifications = Notification.objects.filter(user=request.user, is_read=False)
        count = unread_notifications.update(is_read=True)
        
        return JsonResponse({
            'success': True,
            'marked_as_read': count,
            'message': f'{count} notifications marked as read'
        })
    
    return JsonResponse({'success': False, 'error': 'POST request required'}, status=400)
